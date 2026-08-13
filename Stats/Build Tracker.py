# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from sponsor_data import ROWS, TIERS

wb = Workbook()

# ---------- Styles ----------
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
GOLD = "BF8F00"
WHITE = "FFFFFF"
HEADER_FONT = Font(name="Arial", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="666666")
BODY_FONT = Font(name="Arial", size=10)
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")  # blue = user-editable input
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ================= SHEET 1: Sponsor Tracker =================
ws = wb.active
ws.title = "Sponsor Tracker"

headers = ["Company", "Sector", "Suggested Tier", "Suggested Ask (USD)",
           "Recommended '[What we recommend]' line", "Primary Contact",
           "Contact Type", "Source / Notes", "Status", "Date Contacted", "Follow-up Notes"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER
ws.row_dimensions[1].height = 34

status_options = ["Not Contacted", "Contacted", "Follow-up Sent", "Meeting Scheduled",
                   "Confirmed", "Declined", "No Response"]

for i, row in enumerate(ROWS, start=2):
    if len(row) != 7:
        raise ValueError(f"Invalid row: {row}")
    company, sector, tier, line, contact, ctype, notes = row
    ask = TIERS.get(tier, 0)
    ws.cell(row=i, column=1, value=company)
    ws.cell(row=i, column=2, value=sector)
    ws.cell(row=i, column=3, value=tier)
    ws.cell(row=i, column=4, value=ask)
    ws.cell(row=i, column=5, value=line)
    ws.cell(row=i, column=6, value=contact)
    ws.cell(row=i, column=7, value=ctype)
    ws.cell(row=i, column=8, value=notes)
    ws.cell(row=i, column=9, value="Not Contacted")
    ws.cell(row=i, column=10, value=None)
    ws.cell(row=i, column=11, value=None)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=i, column=col)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    ws.cell(row=i, column=4).number_format = '$#,##0'
    ws.cell(row=i, column=4).alignment = Alignment(horizontal="right", vertical="top")

last_row = len(ROWS) + 1

# Data validation dropdown for Status column
dv = DataValidation(type="list", formula1='"' + ",".join(status_options) + '"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"I2:I{last_row}")

# Column widths
widths = {"A": 30, "B": 26, "C": 24, "D": 14, "E": 55, "F": 30, "G": 26, "H": 40, "I": 16, "J": 14, "K": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"

table = Table(
    displayName="SponsorTracker",
    ref=f"A1:K{last_row}"
)

style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

table.tableStyleInfo = style
ws.add_table(table)

# ================= SHEET 2: Summary =================
ws2 = wb.create_sheet("Summary")
ws2["B2"] = "RTE Robotics - 2026-27 Sponsorship Outreach Summary"
ws2["B2"].font = TITLE_FONT
ws2.merge_cells("B2:D2")

ws2["B4"] = "Season fundraising target (Qatar Championship)"
ws2["B4"].font = BODY_FONT
ws2["C4"] = 11000
ws2["C4"].font = INPUT_FONT
ws2["C4"].number_format = '$#,##0'
ws2["D4"] = "Source: sponsorship prospectus"
ws2["D4"].font = NOTE_FONT

ws2["B5"] = "Total prospects on list"
ws2["B5"].font = BODY_FONT
ws2["C5"] = f"=COUNTA('Sponsor Tracker'!A2:A{last_row})"
ws2["C5"].font = BODY_FONT

ws2["B6"] = "Confirmed pledges (sum)"
ws2["B6"].font = BODY_FONT
ws2["C6"] = f"=SUMIF('Sponsor Tracker'!I2:I{last_row},\"Confirmed\",'Sponsor Tracker'!D2:D{last_row})"
ws2["C6"].font = BODY_FONT
ws2["C6"].number_format = '$#,##0'

ws2["B7"] = "Progress toward target"
ws2["B7"].font = BODY_FONT
ws2["C7"] = "=IFERROR(C6/C4,0)"
ws2["C7"].font = BODY_FONT
ws2["C7"].number_format = "0.0%"

ws2["B9"] = "Prospects by status"
ws2["B9"].font = Font(name="Arial", size=11, bold=True, color=NAVY)
row = 10
for s in status_options:
    ws2.cell(row=row, column=2, value=s).font = BODY_FONT
    ws2.cell(row=row, column=3, value=f'=COUNTIF(\'Sponsor Tracker\'!I2:I{last_row},"{s}")').font = BODY_FONT
    row += 1

ws2["B18"] = "Suggested ask by tier (companies on list)"
ws2["B18"].font = Font(name="Arial", size=11, bold=True, color=NAVY)
row = 19
for tier_name, price in TIERS.items():
    ws2.cell(row=row, column=2, value=tier_name).font = BODY_FONT
    ws2.cell(row=row, column=3, value=f'=COUNTIF(\'Sponsor Tracker\'!C2:C{last_row},"{tier_name}")').font = BODY_FONT
    ws2.cell(row=row, column=4, value=price).font = BODY_FONT
    ws2.cell(row=row, column=4).number_format = '$#,##0'
    row += 1
ws2.cell(row=18, column=3).font = Font(name="Arial", size=9, italic=True, color="666666")
ws2["C18"] = "# companies"
ws2["C18"].font = NOTE_FONT
ws2["D18"] = "tier ask"
ws2["D18"].font = NOTE_FONT

ws2["B26"] = "Legend"
ws2["B26"].font = Font(name="Arial", size=11, bold=True, color=NAVY)
ws2["B27"] = "Contact Type = 'Verified' means a real, sourced email/contact channel was found via research."
ws2["B27"].font = NOTE_FONT
ws2["B28"] = "Contact Type = 'Unverified - use website' means no public CSR/sponsorship email was confirmed;"
ws2["B28"].font = NOTE_FONT
ws2["B29"] = "search the listed domain's Contact Us / CSR / Sustainability page before emailing."
ws2["B29"].font = NOTE_FONT
ws2["B30"] = "Edit the Status, Date Contacted and Follow-up Notes columns on the Sponsor Tracker tab as you go."
ws2["B30"].font = NOTE_FONT

for col, w in {"A": 3, "B": 42, "C": 16, "D": 14}.items():
    ws2.column_dimensions[col].width = w

# ================= SHEET 3: Tier Reference =================
ws3 = wb.create_sheet("Tier Reference")
ws3["B2"] = "2026-27 Sponsorship Tiers (from prospectus)"
ws3["B2"].font = TITLE_FONT
ws3.merge_cells("B2:E2")

tier_info = [
    ("BYTE", 200, 730, "Logo on team shirts; recognition on team social media; name in season-end impact & results report; certificate of appreciation."),
    ("KILOBYTE", 500, 1800, "All BYTE benefits, plus: logo on robot chassis; logo on pit display banner; mid-season performance report; invitation to a team build session or live robot demo."),
    ("MEGABYTE", 1000, 3600, "All KILOBYTE benefits, plus: greater logo prominence on shirts/robot; recognition in Qatar FTC Championship materials; season highlight video (60-90s); featured sponsor mention at outreach events/STEM workshops."),
    ("GIGABYTE", 3000, 10900, "All MEGABYTE benefits, plus: primary branding across all team assets; quarterly technical/impact reports; direct access to team leadership; priority for multi-season partnership."),
    ("TERABYTE (Innovation Sponsor)", 5000, 18200, "All GIGABYTE benefits, plus: subsystem naming rights; pit banner dominance; employee 'Lunch and Learn'; co-authored press release; co-hosted STEM outreach event."),
    ("PETABYTE (Title Sponsor)", 8000, 29100, "All TERABYTE benefits, plus: exclusive 'RTE Robotics - presented by...' naming rights; live office robot demo; dedicated sponsor spotlight video; YouTube build series feature; social media takeover."),
]
headers3 = ["Tier", "USD", "~QAR", "Benefits"]
ws3.append([""] + headers3)
for col in range(2, 6):
    c = ws3.cell(row=3, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER
r = 4
for name, usd, qar, benefits in tier_info:
    ws3.cell(row=r, column=2, value=name).font = Font(name="Arial", size=10, bold=True)
    ws3.cell(row=r, column=3, value=usd).number_format = '$#,##0'
    ws3.cell(row=r, column=4, value=qar).number_format = '#,##0'
    ws3.cell(row=r, column=5, value=benefits)
    for col in range(2, 6):
        cell = ws3.cell(row=r, column=col)
        cell.alignment = WRAP
        cell.border = BORDER
    r += 1

ws3["B12"] = "Season raises toward: $11,000 / ~40,000 QAR (Qatar Championship). Deadline: Nov 1, 2026 (prefer by Kickoff, Sept 12, 2026)."
ws3["B12"].font = NOTE_FONT
ws3.merge_cells("B12:E12")

for col, w in {"A": 3, "B": 26, "C": 12, "D": 12, "E": 70}.items():
    ws3.column_dimensions[col].width = w

wb.save("RTE_Robotics_Sponsor_Tracker.xlsx")
print("Saved.")