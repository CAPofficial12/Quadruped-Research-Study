import pandas as pd
import numpy as np
from itertools import product
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

def generate_expanded_experiment(original_filepath, output_filepath):
    print("=" * 60)
    print("EXPANDING EXPERIMENT DESIGN")
    print("=" * 60)

    # 1. Define the Independent Variables (Factors)
    frequencies = [0.5, 1.0, 1.5, 2.0]
    gaits = ['Walk', 'Trot', 'Pace']
    couplings = [2, 5, 10]
    terrains = [0, 10, 20]
    oscillators = ['Hopf', 'Matsuoka', 'VDP', 'Kuramoto']
    reps = list(range(1, 6)) # 5 repetitions

    # 2. Generate the full 432-condition matrix
    print("Generating 432 unique conditions (108 original x 4 oscillators)...")
    conditions = list(product(frequencies, gaits, couplings, terrains, oscillators))
    condition_matrix = pd.DataFrame(conditions, columns=[
        'Frequency (Hz)', 'Gait', 'Coupling (w)', 'Terrain (deg)', 'Oscillator'
    ])
    condition_matrix.insert(0, 'Condition #', range(1, len(condition_matrix) + 1))

    # 3. Generate the 2160-trial Trial Log
    print("Generating 2,160 randomized trials...")
    trials = []
    trial_id = 1
    for _, row in condition_matrix.iterrows():
        for rep in reps:
            trials.append({
                'Trial ID': trial_id,
                'Condition #': row['Condition #'],
                'Rep #': rep,
                'Frequency (Hz)': row['Frequency (Hz)'],
                'Gait': row['Gait'],
                'Coupling (w)': row['Coupling (w)'],
                'Terrain (deg)': row['Terrain (deg)'],
                'Oscillator': row['Oscillator']
            })
            trial_id += 1

    trial_log = pd.DataFrame(trials)
    
    # Randomize the run order
    trial_log = trial_log.sample(frac=1, random_state=42).reset_index(drop=True)
    trial_log.insert(0, 'Run Order', range(1, len(trial_log) + 1))

    # 4. Define Original and New DV columns
    original_dv_inputs = ['Distance (m)', 'Time (s)', 'RMS Tilt (deg)', 'Fall? (Y/N)', 'Energy (J)', 'Robot Mass (kg)', 'Battery V (start)']
    
    new_dvs = [
        'Avg_Power_W', 'Energy_per_m', 'Vel_Stability', 'CoM_Lat_Osc', 
        'CoM_Vert_Osc', 'Pitch_Amp', 'Roll_Amp', 'ZMP_Dev', 
        'Stability_Margin', 'Stride_Length', 'Duty_Cycle', 'Foot_Slip', 
        'Phase_Error', 'Tracking_Error'
    ]

    # Add placeholders for all inputs/DVs to be filled by the user
    for col in original_dv_inputs + new_dvs:
        trial_log[col] = np.nan

    # Calculate original auto-calculated columns (Speed, COT) just in case they want them in the flat sheet
    # Though they might rely on Excel formulas, we'll leave them blank to be filled or calculated later
    trial_log['Speed (m/s)'] = np.nan
    trial_log['COT'] = np.nan
    trial_log['Notes'] = ""

    # Rearrange columns to match standard structure
    cols_order = [
        'Run Order', 'Trial ID', 'Condition #', 'Rep #', 
        'Frequency (Hz)', 'Gait', 'Coupling (w)', 'Terrain (deg)', 'Oscillator',
        'Distance (m)', 'Time (s)', 'Speed (m/s)', 
        'Avg_Power_W', 'Energy_per_m', 'COT',
        'RMS Tilt (deg)', 'Fall? (Y/N)', 'Vel_Stability', 'CoM_Lat_Osc', 'CoM_Vert_Osc', 
        'Pitch_Amp', 'Roll_Amp', 'ZMP_Dev', 'Stability_Margin',
        'Stride_Length', 'Duty_Cycle', 'Foot_Slip', 'Phase_Error', 'Tracking_Error',
        'Energy (J)', 'Robot Mass (kg)', 'Battery V (start)', 'Notes'
    ]
    trial_log = trial_log[cols_order]

    # 5. Update the Factors & Levels sheet
    factors_data = [
        ['1. Frequency (v) - 4 levels', '', '', ''],
        ['Level', 'Value', 'Notes', ''],
        [1, 0.5, 'Slowest tested stepping rate', ''],
        [2, 1.0, 'Baseline mid-range rate', ''],
        [3, 1.5, 'Fast rate', ''],
        [4, 2.0, 'Practical ceiling - approaches DS3218 Pro swing-phase speed limit', ''],
        ['', '', '', ''],
        ['2. Gait - 3 levels (phase bias, fraction of cycle, FL = reference = 0.00)', '', '', ''],
        ['Gait', 'Front-Left', 'Front-Right', 'Hind-Left', 'Hind-Right', 'Description'],
        ['Walk', 0, 0.5, 0.25, 0.75, 'Statically stable, sequential leg gait'],
        ['Trot', 0, 0.5, 0.5, 0, 'Diagonal leg pairs move together'],
        ['Pace', 0, 0.5, 0, 0.5, 'Same-side leg pairs move together'],
        ['', '', '', ''],
        ['3. Coupling weight (w) - 3 levels', '', '', ''],
        ['Level', 'Value', 'Notes', ''],
        [1, 2, 'Weak coupling - gait pattern can drift/desync under perturbation', ''],
        [2, 5, 'Moderate coupling - comparable to the frequency term at ~1 Hz', ''],
        [3, 10, 'Strong coupling - gait pattern held rigidly against perturbation', ''],
        ['', '', '', ''],
        ['4. Terrain gradient - 3 levels', '', '', ''],
        ['Level', 'Value', 'Notes', ''],
        [1, 0, 'Flat ground baseline', ''],
        [2, 10, 'Moderate incline', ''],
        [3, 20, 'Steep incline - stress test for stability/coupling', ''],
        ['', '', '', ''],
        ['5. Oscillator Type - 4 levels (NEW IV)', '', '', ''],
        ['Level', 'Value', 'Description', ''],
        [1, 'Hopf', 'Smooth, easy frequency modulation, most cited baseline', ''],
        [2, 'Matsuoka', 'Neuron-inspired, harder to tune but biologically motivated', ''],
        [3, 'VDP', 'Van der Pol - strong synchronization, good for straight-line locomotion', ''],
        [4, 'Kuramoto', 'Phase oscillator model, excellent for analyzing phase coupling', ''],
        ['', '', '', ''],
        ['Total Conditions: 4 x 3 x 3 x 3 x 4 = 432 conditions x 5 reps = 2,160 trials', '', '', '']
    ]
    factors_df = pd.DataFrame(factors_data)

    # 6. Write to a new Excel file with formatting
    print(f"Writing expanded design to {output_filepath}...")
    
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        # Write sheets
        factors_df.to_excel(writer, sheet_name='Factors & Levels', index=False, header=False)
        condition_matrix.to_excel(writer, sheet_name='Condition Matrix', index=False)
        trial_log.to_excel(writer, sheet_name='Trial Log', index=False)
        
        # Write a README
        readme_data = {
            'Sheet': ['README', 'Factors & Levels', 'Condition Matrix', 'Trial Log'],
            'Description': [
                'Expanded experiment tracking workbook - 432 conditions x 5 repetitions = 2,160 trials',
                'The 5 independent variables (including NEW Oscillator IV) and their levels.',
                'The 432 unique parameter combinations.',
                'All 2,160 trials in randomized order. Blue cells are inputs you measure.'
            ]
        }
        pd.DataFrame(readme_data).to_excel(writer, sheet_name='README', index=False)

        # Get workbook to apply formatting
        workbook = writer.book
        worksheet = writer.sheets['Trial Log']
        
        # Define fills
        blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Color the input columns (Columns J onwards to AF)
        # J=10 (Distance) to AF=32 (Tracking_Error)
        for col_idx in range(10, 33):
            for row_idx in range(2, len(trial_log) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = blue_fill
                
        # Color IV columns (E to I) lightly to distinguish
        for col_idx in range(5, 10):
             for row_idx in range(2, len(trial_log) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = green_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print("\n" + "=" * 60)
    print("SUCCESS! NEW EXCEL FILE CREATED.")
    print("=" * 60)
    print(f"File saved as: {output_filepath}")
    print(f"Total Trials in Sheet: {len(trial_log)}")
    print("\nNew IV 'Oscillator' added (Column I).")
    print("New DV columns added (Columns N through AF).")

# --- Run the script ---
if __name__ == "__main__":
    # Make sure to replace this with the actual path to your file
    input_file = "CPG_Quadruped_Trial_Log.xlsx"
    output_file = "CPG_Quadruped_Trial_Log_EXPANDED.xlsx"
    
    if os.path.exists(input_file):
        generate_expanded_experiment(input_file, output_file)
    else:
        print(f"Error: Could not find '{input_file}'.")
        print("Please ensure the script is in the same folder as your Excel file, or provide the full path.")